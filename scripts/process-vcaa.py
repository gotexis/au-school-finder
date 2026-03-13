#!/usr/bin/env python3
"""Process VCAA Senior Secondary data into school VCE scores JSON."""
import json
import openpyxl
import re

wb = openpyxl.load_workbook('scripts/vcaa-2025.xlsx')
ws = wb.active

# Find header row (row 12 based on inspection)
headers = [str(c.value or '').strip() for c in ws[12]]
print(f"Headers: {headers[:5]}...")

schools = []
for row in ws.iter_rows(min_row=13, values_only=True):
    name = row[0]
    if not name or not isinstance(name, str):
        continue
    
    locality = str(row[2] or '').strip()
    
    # Parse numeric fields, handling '-', 'I/D', '< 4'
    def parse_num(val):
        if val is None or val == '-' or val == 'I/D' or str(val).startswith('<'):
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None
    
    median_score = parse_num(row[20])  # Median VCE study score
    pct_40_plus = parse_num(row[21])   # % of study scores >= 40
    pct_completion = parse_num(row[15]) # % satisfactory VCE completions
    num_students = parse_num(row[10])   # Number of VCE students
    pct_vtac = parse_num(row[14])       # % applying for tertiary places
    
    schools.append({
        'name': name.strip(),
        'locality': locality,
        'medianStudyScore': median_score,
        'pct40Plus': pct_40_plus,
        'pctCompletion': pct_completion,
        'numVceStudents': int(num_students) if num_students else None,
        'pctVtac': pct_vtac,
    })

# Filter to schools with a median score
with_score = [s for s in schools if s['medianStudyScore'] is not None]
with_score.sort(key=lambda s: s['medianStudyScore'], reverse=True)

print(f"Total rows: {len(schools)}, with median score: {len(with_score)}")
print(f"Top 5: {[(s['name'], s['medianStudyScore']) for s in with_score[:5]]}")

# Save
with open('src/data/vce-scores.json', 'w') as f:
    json.dump(with_score, f, indent=2)

print(f"Saved {len(with_score)} schools to src/data/vce-scores.json")
